"""用截图里的真实数字造一个测试 Excel,完整复刻便宜坊日报的二维布局。

这个脚本只为验证 parser 能不能正确读取你的表格。
你拿到真实 .xlsx 文件后,这个脚本就不需要了。
"""
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "日报"

# ========== 左半:营业收入 ==========
# 标题
ws["A1"] = "便宜坊  马连道  店营业收入日报表    2026 年 5 月 27 日"
ws.merge_cells("A1:D1")
ws["A1"].alignment = Alignment(horizontal="center")
ws["A1"].font = Font(bold=True)

# 左半数据(列 A=字段名,列 B=数值,列 C=字段名,列 D=数值)
left_block = [
    ("本日收入", 20617.78, "日折前营业额", 35120.9),
    ("月累计收入", 880095.34, "月折前营业额累计", 1467605.04),
    ("同期月累计收入", 1153319.84, "冰箱贴发放数量", None),
    ("同比月累计收入差额", -273224.5, "发布笔记赠送数量", 1),
    ("堂食收入", 18476.5, "集章兑换数量", 0),
    ("堂食外卖收入", 330, None, None),
    (None, None, "儿童集章卡", None),
    ("免单金额", 0, "本日发放数量", 0),
    ("进货金额", 16734.06, "月累计发放数量", 35),
    ("堂食会员消费收入", 7059, "堂食会员消费占比", 37.53),
    ("堂食正价消费收入", 3522, "堂食原价消费占比", 18.72),
    ("堂食其他优惠消费收入", 8225.5, "堂食其他优惠消费占比", 43.74),
    (None, None, "发放1416烤鸭券数量", 4),
    ("本日线上外卖收入", 1811.28, "回收1416烤鸭券数量", 1),
    ("月线上外卖累计收入", 107858.86, "烤鸭券带来收入", 660.00),
    (None, None, "回收100元代金券数量", 1),
    ("本日会员储值", 3000, "代金券带来收入", 300.00),
    ("月累计会员储值", 363237.81, "春笋滑溜里脊", 6),
    (None, None, "带来收入", 1644.00),
    ("来客数", 133, None, None),
    ("客单价", 138.9210526, None, None),
]
start_row = 2
for i, row in enumerate(left_block):
    r = start_row + i
    if row[0] is not None:
        ws.cell(row=r, column=1, value=row[0])
    if row[1] is not None:
        ws.cell(row=r, column=2, value=row[1])
    if row[2] is not None:
        ws.cell(row=r, column=3, value=row[2])
    if row[3] is not None:
        ws.cell(row=r, column=4, value=row[3])

# ========== 右半:菜品销售(列 F 开始)==========
# 注意:第一列是大类名(合并单元格),第二列是字段名,第三列是数值
right_blocks = [
    # (大类, [(字段名, 值), ...])
    ("烤鸭+鸭架+烤鸭烧饼销售数据", [
        ("堂食烤鸭日销售数量", 24.5),
        ("迷你烤鸭日销售数量", 6),
        ("线上外卖烤鸭日销售数量", 8),
        ("日累计", 38.5),
        ("月累计", 1648),
        ("椒盐_香辣鸭架日销售数量", 10),
        ("日累计", None),
        ("月累计", 414),
        ("鸭架占比", 76.92),
        ("烤鸭烧饼日销售数量", 72),
        ("烤鸭小料日销售数量", 127),
        ("烧饼占比", 56.69),
    ]),
    ("套餐+乳鸽销售数据", [
        ("3人套餐日销售数量", 0),
        ("6人套餐日销售数量", 0),
        ("8人套餐日销售数量", 0),
        ("10人套餐日销售数量", 0),
        ("12人套餐日销售数量", 0),
        ("松叶蟹套餐日销售数量", 3),
        ("日累计", 3),
        ("月累计", 80),
        ("乳鸽日销售数量", 9),
        ("月累计", 362),
    ]),
    ("鱼类+牛掌销售数据", [
        ("鳜鱼套日销售数量", 2),
        ("鱼类日销售数量", 5),
        ("海参_烧绘牛掌日销售数量", 1),
        ("日累计", 8),
        ("月累计", 410),
    ]),
    ("位吃+甜品+自制销售数据", [
        ("点心日销售数量", 0),
        ("位吃日销售数量", 56),
        ("月累计", 2043),
        ("甜品日销售数量", 16),
        ("月累计", 483),
        ("自制饮品日销售数量", 12),
        ("月累计", 399),
    ]),
    ("精酿销售数据", [
        ("精酿啤酒日销售数量", 6),
        ("月累计", 6),
    ]),
]

# 右半起始列:从 F 列(索引 6)开始
COL_CAT = 6     # 大类名所在列
COL_FIELD = 7   # 字段名所在列
COL_VALUE = 8   # 数值所在列

ws.cell(row=1, column=COL_CAT, value="便宜坊  马连道  店销售日报表   2026 年 5 月 27 日")
ws.merge_cells(start_row=1, start_column=COL_CAT, end_row=1, end_column=COL_VALUE)
ws.cell(row=1, column=COL_CAT).alignment = Alignment(horizontal="center")
ws.cell(row=1, column=COL_CAT).font = Font(bold=True)

cur_row = 2
for cat_name, fields in right_blocks:
    block_start = cur_row
    for fname, fval in fields:
        if fname is not None:
            ws.cell(row=cur_row, column=COL_FIELD, value=fname)
        if fval is not None:
            ws.cell(row=cur_row, column=COL_VALUE, value=fval)
        cur_row += 1
    # 合并大类名(纵向)
    ws.cell(row=block_start, column=COL_CAT, value=cat_name)
    if cur_row - 1 > block_start:
        ws.merge_cells(start_row=block_start, start_column=COL_CAT,
                       end_row=cur_row - 1, end_column=COL_CAT)
    ws.cell(row=block_start, column=COL_CAT).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)
    ws.cell(row=block_start, column=COL_CAT).font = Font(bold=True)

# 列宽
for col_letter, width in [("A", 22), ("B", 12), ("C", 22), ("D", 12),
                           ("F", 15), ("G", 24), ("H", 12)]:
    ws.column_dimensions[col_letter].width = width

# 保存
output_path = Path(__file__).parent / "data" / "便宜坊马连道_2026-05-27.xlsx"
output_path.parent.mkdir(exist_ok=True)
wb.save(output_path)
print(f"✅ 已生成模拟日报: {output_path}")
print(f"   你可以用 'python main.py --file {output_path.name}' 测试解析")
