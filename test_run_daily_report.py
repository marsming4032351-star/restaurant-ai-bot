import csv
import json
import tempfile
import unittest
from datetime import date
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import run_daily_report as R
import watch_daily_folder as W
import parser as P
import openpyxl


class RunDailyReportTests(unittest.TestCase):
    def test_default_input_dir_is_maliandao_service_friendly_folder(self):
        expected = Path("/Users/ming/Restaurant/daily-input/马连道")

        self.assertEqual(R.INPUT_DIR, expected)

    def test_latest_input_image_allows_explicit_input_folder(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            old = folder / "old.png"
            new = folder / "new.jpg"
            ignored = folder / "note.txt"
            old.write_bytes(b"old")
            new.write_bytes(b"new")
            ignored.write_text("ignore", encoding="utf-8")

            image = R.latest_input_image(folder)

            self.assertEqual(image, new)

    def test_extract_json_from_markdown_block(self):
        text = '结果如下：\n```json\n{"本日收入": 123.45, "来客数": 9}\n```'

        self.assertEqual(R.extract_json_object(text), {"本日收入": 123.45, "来客数": 9})

    def test_pipeline_log_upgrade_and_success_append(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "pipeline_log.csv"
            path.write_text(
                "date,store_name,feishu_pushed,status,notes\n"
                "2026-05-28,便宜坊马连道,true,done,\n",
                encoding="utf-8",
            )

            R.append_pipeline_log(
                path,
                {
                    "date": "2026-05-29",
                    "store_name": "便宜坊马连道",
                    "feishu_pushed": "true",
                    "feishu_push_success": "true",
                    "status": "done",
                    "notes": "",
                },
            )

            with path.open(encoding="utf-8", newline="") as f:
                rows = list(csv.DictReader(f))
            self.assertIn("feishu_push_success", rows[0])
            self.assertIn("error_message", rows[0])
            self.assertEqual(rows[-1]["date"], "2026-05-29")
            self.assertEqual(rows[-1]["feishu_push_success"], "true")
            self.assertEqual(rows[-1]["status"], "done")

    def test_is_already_pushed_reads_new_or_old_success_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "pipeline_log.csv"
            path.write_text(
                "date,store_name,feishu_pushed,feishu_push_success,status\n"
                "2026-05-29,便宜坊马连道,true,,done\n"
                "2026-05-30,便宜坊马连道,false,true,done\n",
                encoding="utf-8",
            )

            self.assertTrue(R.is_already_pushed(path, "2026-05-29", "便宜坊马连道"))
            self.assertTrue(R.is_already_pushed(path, "2026-05-30", "便宜坊马连道"))
            self.assertFalse(R.is_already_pushed(path, "2026-05-31", "便宜坊马连道"))

    def test_write_pipeline_state_advances_target_date(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "pipeline_state.json"

            R.write_pipeline_state(path, "便宜坊马连道", "2026-05-29", "2026-05-30T10:00:00+08:00")

            data = json.loads(path.read_text(encoding="utf-8"))
            self.assertEqual(data["current_target_date"], "2026-05-30")
            self.assertEqual(data["last_completed_date"], "2026-05-29")
            self.assertTrue(data["last_feishu_pushed"])
            self.assertEqual(data["updated_by"], "codex")

    def test_store_history_duplicate_detection(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "store_history.csv"
            path.write_text(
                "date,store_name,revenue\n"
                "2026-05-29,便宜坊马连道,35447.22\n",
                encoding="utf-8",
            )

            self.assertTrue(R.store_history_has_row(path, "2026-05-29", "便宜坊马连道"))
            self.assertFalse(R.store_history_has_row(path, "2026-05-30", "便宜坊马连道"))

    def test_image_header_business_date_overrides_processing_date_for_daily_and_weekly(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            image = base / "daily.png"
            image.write_bytes(b"fake image")
            pipeline_state = base / "pipeline_state.json"
            pipeline_log = base / "pipeline_log.csv"
            data_dir = base / "data"
            output_dir = base / "output"
            output_dir.mkdir()
            weekly_calls = []
            main_calls = []

            def fake_build_excel(daily_json, business_date, output_dir_arg):
                self.assertEqual(daily_json["业务日期"], "2026-05-31")
                self.assertEqual(business_date, "2026-05-31")
                path = output_dir_arg / f"便宜坊马连道_{business_date}.xlsx"
                path.parent.mkdir(parents=True, exist_ok=True)
                path.write_text("fake excel", encoding="utf-8")
                return path

            def fake_main_run(report_date, store_id, excel_path, args_ns=None):
                self.assertIsNone(report_date)
                self.assertEqual(store_id, "MLD")
                self.assertEqual(excel_path.name, "便宜坊马连道_2026-05-31.xlsx")
                report_path = output_dir / "report_MLD_2026-05-31.json"
                report_path.write_text(
                    json.dumps(
                        {
                            "daily": {
                                "meta": {
                                    "date": "2026-05-31",
                                    "store_name": "便宜坊马连道",
                                }
                            },
                            "report": {"headline": "ok"},
                        },
                        ensure_ascii=False,
                    ),
                    encoding="utf-8",
                )
                main_calls.append(excel_path)

            args = SimpleNamespace(
                image=str(image),
                input_folder=str(base),
                store="便宜坊马连道",
                store_id="MLD",
                date="2026-06-01",
                force=False,
            )

            with patch.object(R, "PIPELINE_STATE", pipeline_state), \
                 patch.object(R, "PIPELINE_LOG", pipeline_log), \
                 patch.object(R.config, "DATA_DIR", data_dir), \
                 patch.object(R.config, "OUTPUT_DIR", output_dir), \
                 patch.object(R, "read_startup_context", return_value=None), \
                 patch.object(R, "recognize_image", return_value={"业务日期": "2026-05-31", "本日收入": 1}), \
                 patch.object(R.image_to_excel, "build_excel", side_effect=fake_build_excel), \
                 patch.dict("sys.modules", {"main": SimpleNamespace(run=fake_main_run)}), \
                 patch.object(R, "store_history_has_row", return_value=False), \
                 patch.object(R, "run_git_commit_push", return_value=None), \
                 patch.object(R, "now_iso", return_value="2026-06-01T06:00:00+08:00"), \
                 patch.object(R.weekly_auto, "today_in_business_timezone", return_value=date(2026, 6, 1)), \
                 patch.object(R.weekly_auto, "check_and_push", side_effect=lambda store, completed_date: weekly_calls.append((store, completed_date)) or {"triggered": True, "start_date": "2026-05-25", "end_date": "2026-05-31"}):
                exit_code = R.run_daily_report(args)

            self.assertEqual(exit_code, 0)
            self.assertEqual(len(main_calls), 1)
            self.assertEqual(weekly_calls, [("便宜坊马连道", "2026-05-31")])
            state = json.loads(pipeline_state.read_text(encoding="utf-8"))
            self.assertEqual(state["last_completed_date"], "2026-05-31")
            self.assertEqual(state["current_target_date"], "2026-06-01")
            with pipeline_log.open(encoding="utf-8", newline="") as f:
                rows = list(csv.DictReader(f))
            self.assertEqual(rows[-1]["date"], "2026-05-31")
            self.assertIn("便宜坊马连道_2026-05-31.xlsx", rows[-1]["excel_file"])
            self.assertIn("report_MLD_2026-05-31.json", rows[-1]["report_file"])

    def test_parser_prefers_report_header_date_over_supplied_processing_date(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "便宜坊马连道_2026-06-01.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = "便宜坊  马连道  店营业收入日报表    2026 年 5 月 31 日"
            ws["A2"] = "本日收入"
            ws["B2"] = 100
            ws["A21"] = "来客数"
            ws["B21"] = 10
            ws["A22"] = "客单价"
            ws["B22"] = 10
            wb.save(path)

            daily = P.load_daily(path, date(2026, 6, 1))

            self.assertEqual(daily["meta"]["date"], "2026-05-31")
            self.assertEqual(daily["meta"]["weekday"], "Sunday")

    def test_watch_state_skips_same_file_signature(self):
        with tempfile.TemporaryDirectory() as tmp:
            image = Path(tmp) / "daily.png"
            image.write_bytes(b"abc")
            state_path = Path(tmp) / "watch_state.json"

            signature = W.file_signature(image)
            self.assertTrue(W.should_process({}, image, signature))

            state = {}
            W.mark_processed(state, image, signature, "2026-05-30T10:00:00+08:00")
            self.assertFalse(W.should_process(state, image, signature))

    def test_watcher_does_not_pass_system_date_as_report_date_by_default(self):
        with tempfile.TemporaryDirectory() as tmp:
            image = Path(tmp) / "daily.png"
            image.write_bytes(b"abc")

            with patch.object(W.subprocess, "run") as run:
                W.run_report(image, "便宜坊马连道")

            cmd = run.call_args.args[0]
            self.assertNotIn("--date", cmd)
            self.assertIn("--image", cmd)
            self.assertIn(str(image), cmd)

    def test_find_candidate_images_sorted_by_mtime(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            old = folder / "old.png"
            new = folder / "new.webp"
            ignored = folder / "note.txt"
            old.write_bytes(b"old")
            new.write_bytes(b"new")
            ignored.write_text("ignore", encoding="utf-8")

            images = W.find_candidate_images(folder)

            self.assertEqual(images[-1], new)
            self.assertNotIn(ignored, images)

    def test_launchd_scripts_define_expected_service_contract(self):
        install = Path("scripts/install_watcher_launchd.sh").read_text(encoding="utf-8")
        uninstall = Path("scripts/uninstall_watcher_launchd.sh").read_text(encoding="utf-8")
        status = Path("scripts/status_watcher_launchd.sh").read_text(encoding="utf-8")

        self.assertIn("com.restaurant.daily-watcher", install)
        self.assertIn("/usr/bin/python3", install)
        self.assertIn("/Users/ming/Restaurant/restaurant-ai-bot/watch_daily_folder.py", install)
        self.assertIn("/Users/ming/Restaurant/restaurant-ai-bot/logs/watch_daily_folder.log", install)
        self.assertIn("/Users/ming/Restaurant/daily-input/马连道", install)
        self.assertRegex(install, r"mkdir -p .*\$INPUT_DIR")
        self.assertIn("<key>KeepAlive</key>", install)
        self.assertIn("<key>RunAtLoad</key>", install)
        self.assertRegex(install, r"launchctl (bootstrap|load)")
        self.assertRegex(install, r"launchctl (kickstart|start)")

        self.assertIn("launchctl", uninstall)
        self.assertIn("rm -f \"$PLIST_PATH\"", uninstall)
        self.assertIn("watch_daily_folder.py", status)
        self.assertIn("tail -50", status)


if __name__ == "__main__":
    unittest.main()
