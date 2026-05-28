"""
image_to_excel.py — 把从图片中读取的字段数据写入标准日报 Excel。

用法:
    python image_to_excel.py --date 2026-05-28 --json '{"本日收入": 20617.78, ...}'
    python image_to_excel.py --date 2026-05-28 --json-file data.json

JSON 格式:
    扁平键值对，键名与 seed_real_data.py 中的字段名完全一致。
    右半菜品数据中出现多次的"日累计"/"月累计"，请在键名后加大类前缀区分，例如:
        "烤鸭_日累计", "套餐_月累计", "鱼类_日累计"

输出:
    data/便宜坊马连道_YYYY-MM-DD.xlsx
"""

import argparse
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

# ────────────────────────────────────────────────
# 左半：A=字段名, B=数值, C=字段名, D=数值
# None 表示该格留空（保持行位置与原表一致）
# ────────────────────────────────────────────────
LEFT_LAYOUT = [
    ("本日收入",               "C_日折前营业额",            "日折前营业额"),
    ("月累计收入",             "C_月折前营业额累计",         "月折前营业额累计"),
    ("同期月累计收入",         "C_冰箱贴发放数量",           "冰箱贴发放数量"),
    ("同比月累计收入差额",     "C_发布笔记赠送数量",         "发布笔记赠送数量"),
    ("堂食收入",               "C_集章兑换数量",             "集章兑换数量"),
    ("堂食外卖收入",           None,                         None),
    (None,                     "C_儿童集章卡",               "儿童集章卡"),
    ("免单金额",               "C_本日发放数量",             "本日发放数量"),
    ("进货金额",               "C_月累计发放数量",           "月累计发放数量"),
    ("堂食会员消费收入",       "C_堂食会员消费占比",         "堂食会员消费占比"),
    ("堂食正价消费收入",       "C_堂食原价消费占比",         "堂食原价消费占比"),
    ("堂食其他优惠消费收入",   "C_堂食其他优惠消费占比",     "堂食其他优惠消费占比"),
    (None,                     "C_发放1416烤鸭券数量",       "发放1416烤鸭券数量"),
    ("本日线上外卖收入",       "C_回收1416烤鸭券数量",       "回收1416烤鸭券数量"),
    ("月线上外卖累计收入",     "C_烤鸭券带来收入",           "烤鸭券带来收入"),
    (None,                     "C_回收100元代金券数量",       "回收100元代金券数量"),
    ("本日会员储值",           "C_代金券带来收入",           "代金券带来收入"),
    ("月累计会员储值",         "C_春笋滑溜里脊",             "春笋滑溜里脊"),
    (None,                     "C_带来收入",                 "带来收入"),
    ("来客数",                 None,                         None),
    ("客单价",                 None,                         None),
]
# 每行结构：(A列字段名, C列占位符key, C列字段名)
# B列值 = data[A列字段名]，D列值 = data[C列字段名]

# ────────────────────────────────────────────────
# 右半：F=大类(合并), G=字段名, H=数值
# 重复名称用"大类_字段名"前缀作为 JSON key 查找
# ────────────────────────────────────────────────
RIGHT_LAYOUT = [
    ("烤鸭+鸭架+烤鸭烧饼销售数据", [
        "堂食烤鸭日销售数量",
        "迷你烤鸭日销售数量",
        "线上外卖烤鸭日销售数量",
        "烤鸭_日累计",
        "烤鸭_月累计",
        "椒盐_香辣鸭架日销售数量",
        "鸭架_日累计",
        "鸭架_月累计",
        "鸭架占比",
        "烤鸭烧饼日销售数量",
        "烤鸭小料日销售数量",
        "烧饼占比",
    ]),
    ("套餐+乳鸽销售数据", [
        "3人套餐日销售数量",
        "6人套餐日销售数量",
        "8人套餐日销售数量",
        "10人套餐日销售数量",
        "12人套餐日销售数量",
        "松叶蟹套餐日销售数量",
        "套餐_日累计",
        "套餐_月累计",
        "乳鸽日销售数量",
        "乳鸽_月累计",
    ]),
    ("鱼类+牛掌销售数据", [
        "鳜鱼套日销售数量",
        "鱼类日销售数量",
        "海参_烧绘牛掌日销售数量",
        "鱼类_日累计",
        "鱼类_月累计",
    ]),
    ("位吃+甜品+自制销售数据", [
        "点心日销售数量",
        "位吃日销售数量",
        "位吃_月累计",
        "甜品日销售数量",
        "甜品_月累计",
        "自制饮品日销售数量",
        "自制_月累计",
    ]),
    ("精酿销售数据", [
        "精酿啤酒日销售数量",
        "精酿_月累计",
    ]),
]

# 右半 key→显示名（去掉前缀）
def _display_name(key: str) -> str:
    """烤鸭_日累计 → 日累计，普通字段名原样返回。"""
    if "_" in key:
        parts = key.split("_", 1)
        # 只有后缀是"日累计/月累计/月累计/占比"等短词才去前缀
        suffixes = {"日累计", "月累计", "占比"}
        if parts[1] in suffixes:
            return parts[1]
    return key


def build_excel(data: dict, date_str: str, output_dir: Path) -> Path:
    year, month, day = date_str.split("-")
    title_date = f"{year} 年 {int(month)} 月 {int(day)} 日"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "日报"

    center = Alignment(horizontal="center")
    bold = Font(bold=True)

    # ── 左半标题 ──
    ws["A1"] = f"便宜坊  马连道  店营业收入日报表    {title_date}"
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = center
    ws["A1"].font = bold

    # ── 左半数据 ──
    for i, (a_key, _c_placeholder, c_key) in enumerate(LEFT_LAYOUT):
        r = 2 + i
        if a_key:
            ws.cell(row=r, column=1, value=a_key)
            val = data.get(a_key)
            if val is not None:
                ws.cell(row=r, column=2, value=val)
        if c_key:
            ws.cell(row=r, column=3, value=c_key)
            val = data.get(c_key)
            if val is not None:
                ws.cell(row=r, column=4, value=val)

    # ── 右半标题 ──
    COL_CAT, COL_FIELD, COL_VALUE = 6, 7, 8
    ws.cell(row=1, column=COL_CAT,
            value=f"便宜坊  马连道  店销售日报表   {title_date}")
    ws.merge_cells(start_row=1, start_column=COL_CAT,
                   end_row=1, end_column=COL_VALUE)
    ws.cell(row=1, column=COL_CAT).alignment = center
    ws.cell(row=1, column=COL_CAT).font = bold

    # ── 右半数据 ──
    cur_row = 2
    for cat_name, keys in RIGHT_LAYOUT:
        block_start = cur_row
        for key in keys:
            display = _display_name(key)
            ws.cell(row=cur_row, column=COL_FIELD, value=display)
            val = data.get(key)
            if val is not None:
                ws.cell(row=cur_row, column=COL_VALUE, value=val)
            cur_row += 1
        ws.cell(row=block_start, column=COL_CAT, value=cat_name)
        if cur_row - 1 > block_start:
            ws.merge_cells(start_row=block_start, start_column=COL_CAT,
                           end_row=cur_row - 1, end_column=COL_CAT)
        ws.cell(row=block_start, column=COL_CAT).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=block_start, column=COL_CAT).font = bold

    # ── 列宽 ──
    for col_letter, width in [
        ("A", 22), ("B", 12), ("C", 22), ("D", 12),
        ("F", 15), ("G", 24), ("H", 12),
    ]:
        ws.column_dimensions[col_letter].width = width

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"便宜坊马连道_{date_str}.xlsx"
    wb.save(out_path)
    return out_path


def main():
    parser = argparse.ArgumentParser(description="图片数据 → 日报 Excel")
    parser.add_argument("--date", required=True,
                        help="日报日期，格式 YYYY-MM-DD，例如 2026-05-28")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--json", help="内联 JSON 字符串")
    group.add_argument("--json-file", help="JSON 文件路径")
    parser.add_argument("--output-dir", default=None,
                        help="输出目录，默认为脚本同级的 data/")
    args = parser.parse_args()

    if args.json:
        data = json.loads(args.json)
    else:
        with open(args.json_file, encoding="utf-8") as f:
            data = json.load(f)

    output_dir = Path(args.output_dir) if args.output_dir else \
        Path(__file__).parent / "data"

    out_path = build_excel(data, args.date, output_dir)
    print(f"✅ 已生成: {out_path}")


if __name__ == "__main__":
    main()
